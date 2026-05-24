"""
ETL: carga ventas reales de un kiosco desde Excel.

Lee `~/Downloads/ventas kiosko 2024.xlsx` (hoja "Anexar1"), filtra una sucursal,
y la carga bajo un usuario nuevo "Kiosko Demo".

Uso (desde la raíz del repo):
    docker compose exec api python -m scripts.load_kiosko \\
        --excel /data/ventas_kiosko_2024.xlsx \\
        --branch "Puerto Madero" \\
        --email kiosko-demo@atlas.com \\
        --password demo123

Para que el container vea el archivo, montalo o copialo a `backend/`:
    cp ~/Downloads/ventas\\ kiosko\\ 2024.xlsx backend/ventas_kiosko_2024.xlsx
    docker compose exec api python -m scripts.load_kiosko --excel /app/ventas_kiosko_2024.xlsx
"""
import argparse
import os
import sys
import uuid
from datetime import datetime, time, timezone
from typing import Optional

import bcrypt
import openpyxl
from sqlalchemy import create_engine, select, text
from sqlalchemy.orm import Session, sessionmaker

# Asegura que `app.*` sea importable cuando se corre con `python -m scripts.load_kiosko`
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app.models.category import Category
from app.models.product import Product, product_categories
from app.models.sale import Sale
from app.models.sale_item import SaleItem
from app.models.user import User


# ──────────────────────────────────────────────────────────────
#  Mapeos del Excel al modelo
# ──────────────────────────────────────────────────────────────

PAYMENT_METHOD_MAP = {
    "Tarjeta": "card",
    "Efectivo": "cash",
    "Billetera Virtual": "qr",
}

CARD_TYPE_MAP = {
    "Crédito": "credit",
    "Débito": "debit",
}


def normalize_card_brand(raw: Optional[str]) -> Optional[str]:
    """Visa Débito / Maestro / etc → enum del modelo (visa/mastercard/amex)."""
    if not raw:
        return None
    low = raw.lower()
    if "visa" in low:
        return "visa"
    if "master" in low or "maestro" in low:
        return "mastercard"
    if "amex" in low:
        return "amex"
    return None


# ──────────────────────────────────────────────────────────────
#  Helpers
# ──────────────────────────────────────────────────────────────

def ensure_enum_has_cash(session: Session) -> None:
    """Agrega 'cash' al enum payment_method_enum si no está (Postgres)."""
    try:
        session.execute(text(
            "ALTER TYPE payment_method_enum ADD VALUE IF NOT EXISTS 'cash'"
        ))
        session.commit()
        print("✓ enum payment_method_enum incluye 'cash'")
    except Exception as e:
        session.rollback()
        print(f"⚠ no se pudo ALTER TYPE (puede que ya esté o no sea Postgres): {e}")


def ensure_user(session: Session, email: str, password: str, store_name: str) -> User:
    user = session.execute(select(User).where(User.email == email)).scalar_one_or_none()
    if user:
        print(f"✓ usuario {email} ya existe (id={user.id})")
        return user
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    user = User(
        store_name=store_name,
        email=email,
        password=hashed,
        monthly_goal=500000.0,  # arbitrario, ajustable después
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    print(f"✓ usuario {email} creado (id={user.id})")
    return user


def combine_datetime(d: Optional[datetime], t: Optional[time]) -> datetime:
    base = d or datetime.now(timezone.utc)
    if isinstance(t, time):
        return datetime(base.year, base.month, base.day, t.hour, t.minute, t.second, tzinfo=timezone.utc)
    return base.replace(tzinfo=timezone.utc) if base.tzinfo is None else base


# ──────────────────────────────────────────────────────────────
#  Carga principal
# ──────────────────────────────────────────────────────────────

def load(excel_path: str, branch: str, email: str, password: str, store_name: str,
         max_rows: Optional[int], chunk_size: int) -> None:

    db_url = os.environ.get("DATABASE_URL", "postgresql+psycopg2://postgres:postgres@postgres:5432/postgres")
    engine = create_engine(db_url)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()

    ensure_enum_has_cash(session)
    user = ensure_user(session, email, password, store_name)

    print(f"\nLeyendo Excel: {excel_path}")
    print(f"Sucursal a filtrar: {branch}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Anexar1"]

    # ── Pasada 1: catálogo (categorías y productos únicos del branch) ──
    print("\nPasada 1/2: catálogo")
    cat_names: set[str] = set()
    products_map: dict[int, tuple[str, float, str]] = {}  # ID Producto → (nombre, precio, categoría)

    rows_seen = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1] != branch:
            continue
        prod_id = row[6]
        if prod_id and prod_id not in products_map:
            products_map[prod_id] = (str(row[9]), float(row[11] or 0), str(row[7]))
        if row[7]:
            cat_names.add(str(row[7]))
        rows_seen += 1
        if max_rows and rows_seen >= max_rows:
            break

    print(f"  {len(cat_names)} categorías · {len(products_map)} productos únicos · {rows_seen:,} filas")

    # Upsert categorías
    existing_cats = {
        c.name: c for c in session.execute(
            select(Category).where(Category.user_id == user.id)
        ).scalars().all()
    }
    cat_id_by_name: dict[str, uuid.UUID] = {}
    for name in cat_names:
        if name in existing_cats:
            cat_id_by_name[name] = existing_cats[name].id
        else:
            cat = Category(user_id=user.id, name=name)
            session.add(cat)
            session.flush()
            cat_id_by_name[name] = cat.id
    session.commit()
    print(f"  ✓ {len(cat_id_by_name)} categorías aseguradas")

    # Upsert productos
    existing_prods = {
        p.id: p for p in session.execute(
            select(Product).where(Product.user_id == user.id)
        ).scalars().all()
    }
    excel_to_db_pid: dict[int, uuid.UUID] = {}
    new_products: list[Product] = []
    for excel_pid, (name, price, cat_name) in products_map.items():
        db_pid = uuid.uuid4()
        new_products.append(Product(id=db_pid, user_id=user.id, name=name, price=price))
        excel_to_db_pid[excel_pid] = db_pid

    session.bulk_save_objects(new_products)
    session.commit()

    # Asociar productos a categorías (tabla many-to-many)
    pc_rows = []
    for excel_pid, (_, _, cat_name) in products_map.items():
        pc_rows.append({"product_id": excel_to_db_pid[excel_pid], "category_id": cat_id_by_name[cat_name]})
    if pc_rows:
        session.execute(product_categories.insert(), pc_rows)
        session.commit()
    print(f"  ✓ {len(new_products)} productos insertados y categorizados")

    # ── Pasada 2: ventas (Sale + SaleItem) ──
    print("\nPasada 2/2: ventas")
    sales_buffer: list[dict] = []
    items_buffer: list[dict] = []
    total_inserted = 0
    rows_seen = 0

    def flush():
        nonlocal sales_buffer, items_buffer, total_inserted
        if not sales_buffer:
            return
        session.execute(Sale.__table__.insert(), sales_buffer)
        session.execute(SaleItem.__table__.insert(), items_buffer)
        session.commit()
        total_inserted += len(sales_buffer)
        print(f"  · {total_inserted:,} ventas cargadas")
        sales_buffer = []
        items_buffer = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1] != branch:
            continue

        excel_pid = row[6]
        db_pid = excel_to_db_pid.get(excel_pid)
        if not db_pid:
            continue

        invoice = str(row[3])
        sold_at = combine_datetime(row[4], row[5])
        qty = int(row[12] or 1)
        subtotal = float(row[14] or 0)

        payment = PAYMENT_METHOD_MAP.get(str(row[16]), "cash")
        card_type = CARD_TYPE_MAP.get(row[17]) if row[17] else None
        card_brand = normalize_card_brand(row[18])

        sale_id = uuid.uuid4()
        sales_buffer.append({
            "id": sale_id,
            "user_id": user.id,
            "invoice_number": invoice,
            "payment_method": payment,
            "card_type": card_type,
            "card_brand": card_brand,
            "card_category": None,
            "customer_email": None,
            "clover_order_id": None,
            "total": subtotal,
            "sold_at": sold_at,
        })
        items_buffer.append({
            "id": uuid.uuid4(),
            "sale_id": sale_id,
            "product_id": db_pid,
            "quantity": qty,
            "subtotal": subtotal,
        })

        rows_seen += 1
        if len(sales_buffer) >= chunk_size:
            flush()
        if max_rows and rows_seen >= max_rows:
            break

    flush()
    print(f"\n✓ FIN. {total_inserted:,} ventas cargadas bajo {email}")
    session.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="Ruta al .xlsx dentro del container (ej. /app/ventas_kiosko_2024.xlsx)")
    parser.add_argument("--branch", default="Puerto Madero", help="Nombre del barrio/sucursal a cargar")
    parser.add_argument("--email", default="kiosko-demo@atlas.com")
    parser.add_argument("--password", default="demo123")
    parser.add_argument("--store-name", default="Kiosko Puerto Madero")
    parser.add_argument("--max-rows", type=int, default=None, help="Para pruebas: limita la cantidad de filas")
    parser.add_argument("--chunk-size", type=int, default=2000, help="Filas por commit")
    args = parser.parse_args()

    load(
        excel_path=args.excel,
        branch=args.branch,
        email=args.email,
        password=args.password,
        store_name=args.store_name,
        max_rows=args.max_rows,
        chunk_size=args.chunk_size,
    )
